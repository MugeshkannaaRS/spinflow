"""Smart column mapper for SpinFlow imports.

Maps arbitrary Excel/CSV column headers → SpinFlow system field names
using alias matching + substring scoring + fuzzy Levenshtein fallback.
"""
from __future__ import annotations
import re
import logging
from typing import Any, Optional

logger = logging.getLogger(__name__)


class SmartColumnMapper:
    """
    Maps arbitrary Excel/CSV column headers to SpinFlow system fields.
    """

    def suggest_mapping(
        self,
        headers: list[str],
        module: str,
    ) -> dict[str, str]:
        """
        Returns: {excel_column_header: system_field_name | "custom:HEADER"}

        Priority order:
          1. Exact alias match (score 100)
          2. Substring alias match (score 70)
          3. Word-overlap scoring (score 40-70)
          4. Falls through to "custom:HEADER"
        """
        from app.core.field_aliases import FIELD_ALIASES_BY_MODULE
        aliases = FIELD_ALIASES_BY_MODULE.get(module, {})
        result: dict[str, str] = {}
        used: set[str] = set()

        for header in headers:
            normalized = self._norm(header)
            matched_field: Optional[str] = None
            best_score = 0

            for system_field, alias_list in aliases.items():
                if system_field in used:
                    continue

                for alias in alias_list:
                    na = self._norm(alias)

                    # Exact match
                    if normalized == na:
                        matched_field = system_field
                        best_score = 100
                        break

                    # Contains match (header contains alias or alias contains header)
                    if (na in normalized or normalized in na) and len(na) >= 3:
                        score = 75
                        if score > best_score:
                            best_score = score
                            matched_field = system_field

                    # Word overlap
                    h_words = set(normalized.split())
                    a_words = set(na.split())
                    overlap = len(h_words & a_words)
                    if overlap > 0:
                        score = 40 + (overlap * 15)
                        if score > best_score:
                            best_score = score
                            matched_field = system_field

                if best_score == 100:
                    break

            if matched_field and best_score >= 60:
                result[header] = matched_field
                used.add(matched_field)
            else:
                result[header] = f"custom:{header}"

        return result

    def infer_field_from_values(self, values: list[Any]) -> Optional[str]:
        """Infer system field from sample values."""
        sample = [str(v).strip() for v in values[:10] if v is not None and str(v).strip()][:5]
        if not sample:
            return None

        # Date pattern
        date_pat = re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$")
        if sum(1 for s in sample if date_pat.match(s)) >= 3:
            return "joining_date"

        # Pure numeric
        def to_float(s: str) -> Optional[float]:
            try:
                return float(s.replace(",", ""))
            except ValueError:
                return None

        nums = [to_float(s) for s in sample]
        if all(n is not None for n in nums):
            avg = sum(n for n in nums if n is not None) / len(nums)  # type: ignore[arg-type]
            if 1970 <= avg <= 2035:
                return "manufacturing_year"
            if avg > 5000:
                return "basic"
            return None

        # Gender
        gender_vals = {"m", "f", "male", "female", "m/f", "male/female"}
        if all(s.lower() in gender_vals for s in sample):
            return "gender"

        return None

    def auto_generate_code(self, module: str, row_number: int) -> str:
        """Generate a fallback code when the code column is absent/empty."""
        prefix_map = {
            "machines": "MC",
            "employees": "EMP",
            "departments": "DEPT",
            "customers": "CUST",
            "vehicles": "VEH",
            "routes": "RTE",
            "yarn_counts": "YC",
        }
        prefix = prefix_map.get(module, "REC")
        return f"{prefix}{row_number:04d}"

    @staticmethod
    def _norm(s: str) -> str:
        return s.strip().lower().replace("_", " ").replace("-", " ")


# Module-level singleton
mapper = SmartColumnMapper()


def parse_excel_all_sheets(file_bytes: bytes) -> dict:
    """
    Parse ALL sheets in an Excel workbook, combine into a single row list.
    Header detection: finds the first row with ≥40% non-empty cells and ≥2 string cells.
    Skips annotation rows (< 2 filled cells).
    Returns { headers, rows, total, sheets }.
    """
    import io
    import openpyxl

    if len(file_bytes) > 10 * 1024 * 1024:
        raise ValueError(f"File too large: {len(file_bytes)} bytes (max 10MB)")

    wb = openpyxl.load_workbook(
        filename=io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )

    all_rows: list = []
    detected_headers: Optional[list] = None
    sheet_info: list = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row (first 6 rows)
        header_row_idx: Optional[int] = None
        for idx, row in enumerate(rows[:6]):
            non_empty = [c for c in row if c is not None and str(c).strip()]
            str_cells = [c for c in non_empty if isinstance(c, str)]
            if (len(non_empty) >= max(3, len(row) * 0.4)) and len(str_cells) >= 2:
                header_row_idx = idx
                break

        if header_row_idx is None:
            continue

        raw_headers = rows[header_row_idx]
        headers = [
            str(c).strip() if (c is not None and str(c).strip()) else f"__col_{i}__"
            for i, c in enumerate(raw_headers)
        ]
        # Trim trailing empty columns
        while headers and headers[-1].startswith("__col_"):
            headers.pop()

        if detected_headers is None:
            detected_headers = headers

        sheet_rows = 0
        for row in rows[header_row_idx + 1:]:
            # Skip fully empty rows
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            # Skip rows with < 2 filled cells (section headers / totals)
            filled = [c for c in row if c is not None and str(c).strip() != ""]
            if len(filled) < 2:
                continue

            row_dict: dict = {}
            for i, h in enumerate(headers):
                if i < len(row) and row[i] is not None:
                    row_dict[h] = row[i]

            all_rows.append(row_dict)
            sheet_rows += 1

        sheet_info.append({"sheet": sheet_name, "rows": sheet_rows})

    wb.close()
    return {
        "headers": detected_headers or [],
        "rows": all_rows,
        "total": len(all_rows),
        "sheets": sheet_info,
    }
