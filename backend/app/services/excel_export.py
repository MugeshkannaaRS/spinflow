from __future__ import annotations
from io import BytesIO
from typing import Optional
from openpyxl import Workbook
from app.core.column_labels import get_column_label
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _style_sheet(ws, headers: list[str], rows: list[list]):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def production_report(data: list[dict], module: str = "", field_labels: Optional[dict[str, str]] = None) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Report"

    if not data:
        ws.cell(row=1, column=1, value="No data")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    raw_headers = list(data[0].keys())
    headers = [
        get_column_label(module, h, field_labels) or h
        for h in raw_headers
    ]
    rows = [[str(row.get(h, "")) for h in raw_headers] for row in data]
    _style_sheet(ws, headers, rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def payroll_report(data: list[dict], module: str = "", field_labels: Optional[dict[str, str]] = None) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"

    if not data:
        ws.cell(row=1, column=1, value="No data")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    raw_headers = list(data[0].keys())
    headers = [
        get_column_label(module, h, field_labels) or h
        for h in raw_headers
    ]
    rows = [[str(row.get(h, "")) for h in raw_headers] for row in data]
    _style_sheet(ws, headers, rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gst_report(
    output_gst: dict,
    input_gst: dict,
    net_payable: float,
    month: int,
    year: int,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"GST Summary {month:02d}-{year}"

    rows = [
        ["Component", "Amount (₹)"],
        ["Output CGST", output_gst.get("cgst", 0)],
        ["Output SGST", output_gst.get("sgst", 0)],
        ["Output IGST", output_gst.get("igst", 0)],
        ["Output Total", output_gst.get("total", 0)],
        ["Input GST Total", input_gst.get("total", 0)],
        ["Net Payable", net_payable],
    ]
    _style_sheet(ws, rows[0], [r[1:] for r in rows[1:]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
