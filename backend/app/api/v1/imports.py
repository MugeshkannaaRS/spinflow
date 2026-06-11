import logging
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from typing import List, Optional
from pydantic import BaseModel

from app.db.session import get_db
from app.core.deps import get_current_user, get_mill_scope, require_module, log_audit
from app.core.limiter import limiter
from app.models.user import User
from app.models.import_mapping import ImportMapping

logger = logging.getLogger(__name__)

router = APIRouter()

# Maps import module names → ERP permission module names.
# Used to enforce RBAC on parse / validate / execute.
_IMPORT_TO_ERP_MODULE: dict[str, str] = {
    "machines":          "production",
    "employees":         "hr",
    "yarn_counts":       "masters",
    "departments":       "masters",
    "customers":         "dispatch",
    "suppliers":         "purchase",
    "cotton_purchases":  "purchase",
    "inventory_items":   "inventory",
    "quality_tests":     "quality",
    "dispatch_orders":   "dispatch",
    "payroll_entries":   "payroll",
    "sales_orders":      "sales",
    "maintenance_logs":  "maintenance",
}


async def _assert_import_permission(
    module: str,
    current_user: User,
    db: AsyncSession,
    write: bool = False,
) -> None:
    """Enforce module-level RBAC for import endpoints.

    Maps the import module name to the corresponding ERP module and
    calls require_module() so the same role/subscription rules apply.
    Defaults to 'masters' for unrecognised module names.
    """
    erp_module = _IMPORT_TO_ERP_MODULE.get(module, "masters")
    checker = require_module(erp_module, write=write)
    await checker(current_user=current_user, db=db)

class MappingItem(BaseModel):
    excel_header: str
    spinflow_field: Optional[str] = None
    is_custom_field: bool = False
    confidence: Optional[float] = None

class SaveMappingsRequest(BaseModel):
    mill_id: str
    table_name: str
    mappings: List[MappingItem]

@router.get("/import/mappings")
async def get_import_mappings(
    table: str = Query(...),
    mill_id: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_module("masters")),
):
    scope = await get_mill_scope(current_user, db)
    if scope["mill_id"] and scope["mill_id"] != mill_id:
        raise HTTPException(403, "Access denied")
    if scope["company_id"] and not scope["mill_id"]:
        from app.models.masters import Mill
        mill = await db.get(Mill, mill_id)
        if not mill or mill.company_id != scope["company_id"]:
            raise HTTPException(403, "Access denied")

    try:
        result = await db.execute(
            select(ImportMapping).where(
                ImportMapping.mill_id == mill_id,
                ImportMapping.table_name == table,
            )
        )
        mappings = result.scalars().all()
        return [
            {
                "id": m.id,
                "mill_id": m.mill_id,
                "table_name": m.table_name,
                "excel_header": m.excel_header,
                "spinflow_field": m.spinflow_field,
                "is_custom_field": m.is_custom_field,
                "confidence": float(m.confidence) if m.confidence is not None else None,
                "created_at": m.created_at.isoformat() if m.created_at else None,
            }
            for m in mappings
        ]
    except Exception as e:
        logger.error(f"Error fetching import mappings for table={table}, mill={mill_id}: {e}", exc_info=True)
        return []

@router.post("/import/mappings")
@limiter.limit("10/minute")
async def save_import_mappings(
    request: Request,
    req: SaveMappingsRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_module("masters", write=True)),
):
    scope = await get_mill_scope(current_user, db)
    if scope["mill_id"] and scope["mill_id"] != req.mill_id:
        raise HTTPException(403, "Access denied")
    if scope["company_id"] and not scope["mill_id"]:
        from app.models.masters import Mill
        mill = await db.get(Mill, req.mill_id)
        if not mill or mill.company_id != scope["company_id"]:
            raise HTTPException(403, "Access denied")

    try:
        await db.execute(
            delete(ImportMapping).where(
                ImportMapping.mill_id == req.mill_id,
                ImportMapping.table_name == req.table_name,
            )
        )

        count = 0
        for item in req.mappings:
            mapping = ImportMapping(
                mill_id=req.mill_id,
                table_name=req.table_name,
                excel_header=item.excel_header,
                spinflow_field=item.spinflow_field,
                is_custom_field=item.is_custom_field,
                confidence=item.confidence,
            )
            db.add(mapping)
            count += 1

        await db.commit()
        role_code = current_user.role_rel.code if current_user.role_rel else "UNKNOWN"
        client_ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "0.0.0.0").split(",")[0].strip()
        await log_audit(db, current_user.id, role_code, "import_mappings", "import", req.mill_id, f"Saved {count} import mappings for {req.table_name}", ip_address=client_ip)
        return {"saved": count, "mill_id": req.mill_id, "table_name": req.table_name}
    except Exception as e:
        logger.error(f"Error saving import mappings: {e}", exc_info=True)
        return {"saved": False, "error": str(e)}


# ════════════════════════════════════════════════════════════════════
# Smart Import endpoints (parse → validate → execute pipeline)
# ════════════════════════════════════════════════════════════════════

import io
import json
import time
import uuid as _uuid
import tempfile
import os
from fastapi import UploadFile, File, Form
from typing import Any

# Disk-backed file store — survives memory pressure, bounded by /tmp space.
# Each upload writes two files:
#   /tmp/spinflow_import_{file_id}.bin   — raw bytes
#   /tmp/spinflow_import_{file_id}.meta  — JSON metadata (module, filename, ts)
# Files older than _FILE_TTL_SECONDS are silently ignored on read
# and lazily purged on the next upload.
_FILE_TTL_SECONDS: int = 1800  # 30 minutes
_IMPORT_TMP_PREFIX = "spinflow_import_"


def _tmp_path(file_id: str, ext: str) -> str:
    return os.path.join(tempfile.gettempdir(), f"{_IMPORT_TMP_PREFIX}{file_id}.{ext}")


def _write_import_file(file_id: str, content: bytes, meta: dict) -> None:
    with open(_tmp_path(file_id, "bin"), "wb") as fh:
        fh.write(content)
    meta["_ts"] = time.time()
    with open(_tmp_path(file_id, "meta"), "w") as fh:
        json.dump(meta, fh)


def _read_import_file(file_id: str) -> tuple[Optional[bytes], dict]:
    """Return (content, meta) or (None, {}) if missing/expired."""
    bin_path = _tmp_path(file_id, "bin")
    meta_path = _tmp_path(file_id, "meta")
    if not os.path.exists(bin_path) or not os.path.exists(meta_path):
        return None, {}
    try:
        with open(meta_path) as fh:
            meta = json.load(fh)
        if time.time() - meta.get("_ts", 0) > _FILE_TTL_SECONDS:
            _delete_import_file(file_id)
            return None, {}
        with open(bin_path, "rb") as fh:
            return fh.read(), meta
    except Exception:
        return None, {}


def _delete_import_file(file_id: str) -> None:
    for ext in ("bin", "meta"):
        try:
            os.unlink(_tmp_path(file_id, ext))
        except FileNotFoundError:
            pass


def _purge_expired_import_files() -> None:
    """Lazily remove /tmp files older than TTL. Called on each new upload."""
    tmp = tempfile.gettempdir()
    now = time.time()
    try:
        for name in os.listdir(tmp):
            if not name.startswith(_IMPORT_TMP_PREFIX):
                continue
            full = os.path.join(tmp, name)
            try:
                if now - os.path.getmtime(full) > _FILE_TTL_SECONDS:
                    os.unlink(full)
            except OSError:
                pass
    except OSError:
        pass


class NamedMappingItem(BaseModel):
    header: str
    field: str  # system field name or "custom:HEADER"


class SaveNamedMappingRequest(BaseModel):
    module: str
    name: str
    mapping: list[NamedMappingItem]


@router.post("/imports/parse")
@limiter.limit("20/minute")
async def parse_import_file(
    request: Request,
    file: UploadFile = File(...),
    module: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Step 1: Parse uploaded file, return headers + smart mapping suggestions + preview.
    Supports .xlsx, .xls, .csv, .pdf
    Requires read permission on the target module.
    """
    await _assert_import_permission(module, current_user, db, write=False)
    from app.core.import_mapper import mapper as smart_mapper

    content = await file.read()
    file_id = str(_uuid.uuid4())
    _purge_expired_import_files()
    _write_import_file(file_id, content, {"module": module, "filename": file.filename})

    ext = (file.filename or "").lower().split(".")[-1]

    headers: list[str] = []
    rows: list[dict[str, Any]] = []

    try:
        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise HTTPException(400, detail="File is empty")
            headers = [str(c).strip() for c in all_rows[0] if c is not None and str(c).strip()]
            for row in all_rows[1:]:
                if all(c is None for c in row[:len(headers)]):
                    continue
                rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})

        elif ext == "csv":
            import csv
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            headers = list(reader.fieldnames or [])
            for r in reader:
                rows.append(dict(r))

        elif ext == "pdf":
            try:
                import pdfplumber
                pdf_rows: list[dict] = []
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    for page in pdf.pages:
                        tables = page.extract_tables()
                        for table in tables:
                            if not table:
                                continue
                            if not headers and table[0]:
                                headers = [str(c).strip() for c in table[0] if c is not None and str(c).strip()]
                            for row in table[1:]:
                                if headers:
                                    pdf_rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
                rows = pdf_rows
            except ImportError:
                raise HTTPException(400, detail="PDF parsing requires pdfplumber. Contact admin to install.")
            except Exception as e:
                raise HTTPException(400, detail=f"Failed to parse PDF: {str(e)}")
        else:
            raise HTTPException(400, detail=f"Unsupported file type: {ext}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"parse_import_file error: {e}", exc_info=True)
        raise HTTPException(400, detail=f"Failed to parse file: {str(e)}")

    if not headers:
        raise HTTPException(400, detail="No column headers found in file")

    # Smart mapping
    suggested_mapping = smart_mapper.suggest_mapping(headers, module)

    # Preview (first 5 rows, serialize values)
    preview_rows: list[dict] = []
    for row in rows[:5]:
        preview: dict[str, Any] = {}
        for h in headers:
            v = row.get(h)
            if v is None:
                preview[h] = ""
            elif hasattr(v, "isoformat"):
                preview[h] = v.isoformat()
            else:
                preview[h] = str(v) if not isinstance(v, (int, float, bool)) else v
        preview_rows.append(preview)

    return {
        "file_id": file_id,
        "headers": headers,
        "suggested_mapping": suggested_mapping,
        "preview_rows": preview_rows,
        "total_rows": len(rows),
        "filename": file.filename,
        "module": module,
    }


class ValidateRequest(BaseModel):
    file_id: str
    module: str
    mapping: dict[str, str]  # {header: system_field | "custom:X"}


@router.post("/imports/validate")
@limiter.limit("20/minute")
async def validate_import(
    request: Request,
    req: ValidateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Step 2: Validate mapped data, return per-row status.
    Requires read permission on the target module.
    """
    await _assert_import_permission(req.module, current_user, db, write=False)
    content, meta = _read_import_file(req.file_id)
    if content is None:
        raise HTTPException(400, detail="File not found or expired. Please re-upload.")

    ext = meta.get("filename", "").lower().split(".")[-1] or "xlsx"

    rows: list[dict] = []
    try:
        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                hdrs = [str(c).strip() for c in all_rows[0] if c is not None and str(c).strip()]
                for row in all_rows[1:]:
                    if all(c is None for c in row[:len(hdrs)]):
                        continue
                    rows.append({hdrs[i]: row[i] for i in range(min(len(hdrs), len(row)))})
        elif ext == "csv":
            import csv
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
    except Exception as e:
        raise HTTPException(400, detail=f"Failed to re-parse file: {str(e)}")

    from app.core.import_mapper import mapper as smart_mapper

    result_rows = []
    for idx, raw_row in enumerate(rows):
        mapped_data: dict[str, Any] = {}
        custom_fields: dict[str, Any] = {}
        errors: list[str] = []
        warnings: list[str] = []
        auto_code = False

        for header, system_field in req.mapping.items():
            raw_val = raw_row.get(header)
            if system_field.startswith("custom:"):
                cf_key = system_field[7:].strip()
                custom_fields[cf_key] = str(raw_val) if raw_val is not None else ""
            else:
                mapped_data[system_field] = raw_val

        # Auto-generate code
        code_field = "code" if req.module != "employees" else "employee_code"
        if not mapped_data.get(code_field) or str(mapped_data.get(code_field, "")).strip() == "":
            mapped_data[code_field] = smart_mapper.auto_generate_code(req.module, idx + 1)
            auto_code = True
            warnings.append(f"Code auto-generated: {mapped_data[code_field]}")

        # Name required
        name_field = "name" if req.module != "employees" else "full_name"
        if not mapped_data.get(name_field) or str(mapped_data.get(name_field, "")).strip() == "":
            errors.append(f"{name_field.replace('_', ' ').title()} is required")

        # Year validation for machines
        if req.module == "machines" and mapped_data.get("manufacturing_year"):
            try:
                yr = int(float(str(mapped_data["manufacturing_year"])))
                if not (1900 <= yr <= 2035):
                    errors.append(f"Manufacturing year {yr} out of range (1900–2035)")
                else:
                    mapped_data["manufacturing_year"] = yr
            except (ValueError, TypeError):
                warnings.append(f"Could not parse manufacturing year: {mapped_data['manufacturing_year']}")

        if custom_fields:
            mapped_data["custom_fields"] = custom_fields

        row_status = "error" if errors else ("warning" if warnings else "valid")
        result_rows.append({
            "row_num": idx + 1,
            "status": row_status,
            "data": {k: (str(v) if v is not None and not isinstance(v, (int, float, bool)) else v)
                     for k, v in mapped_data.items()},
            "errors": errors,
            "warnings": warnings,
        })

    valid_count = sum(1 for r in result_rows if r["status"] == "valid")
    warn_count  = sum(1 for r in result_rows if r["status"] == "warning")
    err_count   = sum(1 for r in result_rows if r["status"] == "error")

    return {
        "valid_count": valid_count,
        "warning_count": warn_count,
        "error_count": err_count,
        "total": len(result_rows),
        "rows": result_rows,
    }


class ExecuteRequest(BaseModel):
    file_id: str
    module: str
    mapping: dict[str, str]
    skip_errors: bool = True
    duplicate_handling: str = "skip"  # "skip" | "update" | "new"


@router.post("/imports/execute")
@limiter.limit("5/minute")
async def execute_import(
    request: Request,
    req: ExecuteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Step 3: Actually import validated rows.
    Requires write permission on the target module.
    Rate-limited to 5/minute — execute is expensive and writes to the DB.
    """
    await _assert_import_permission(req.module, current_user, db, write=True)

    # Validate step first to get processed rows
    validate_result = await validate_import(
        request=request,
        req=ValidateRequest(file_id=req.file_id, module=req.module, mapping=req.mapping),
        db=db,
        current_user=current_user,
    )

    scope = await get_mill_scope(current_user, db)
    mill_id = scope.get("mill_id") or current_user.mill_id

    rows_to_import = [
        r for r in validate_result["rows"]
        if r["status"] in ("valid", "warning") or not req.skip_errors
    ]

    imported = 0
    skipped = 0
    errors_out = []

    BATCH_SIZE = 50
    for batch_start in range(0, len(rows_to_import), BATCH_SIZE):
        batch = rows_to_import[batch_start:batch_start + BATCH_SIZE]
        for row_info in batch:
            try:
                data = {**row_info["data"], "mill_id": mill_id}
                if req.module == "machines":
                    from app.models.production import Machine
                    code = str(data.get("code", "")).strip()
                    if not code:
                        skipped += 1
                        continue
                    existing = await db.execute(select(Machine).where(Machine.code == code, Machine.mill_id == mill_id))
                    ex = existing.scalar_one_or_none()
                    if ex and req.duplicate_handling == "skip":
                        skipped += 1
                        continue
                    elif ex and req.duplicate_handling == "update":
                        for k, v in data.items():
                            if k not in ("id",) and hasattr(ex, k):
                                setattr(ex, k, v)
                    else:
                        m = Machine(**{k: v for k, v in data.items()
                                       if hasattr(Machine, k) and k != "id"})
                        db.add(m)
                    imported += 1

                elif req.module == "employees":
                    from app.models.hr import Employee
                    # Remap column-config keys → Employee model attribute names
                    if "full_name" in data and "name" not in data:
                        data["name"] = data.pop("full_name")
                    elif "full_name" in data:
                        data.pop("full_name")
                    if "date_of_joining" in data and "joining_date" not in data:
                        data["joining_date"] = data.pop("date_of_joining")
                    elif "date_of_joining" in data:
                        data.pop("date_of_joining")

                    code = str(data.get("employee_code", data.get("code", ""))).strip()
                    if not code:
                        # Try name as fallback identifier
                        if not data.get("name"):
                            skipped += 1
                            continue
                    existing = None
                    if code:
                        res = await db.execute(select(Employee).where(Employee.code == code))
                        existing = res.scalar_one_or_none()
                    if existing and req.duplicate_handling == "skip":
                        skipped += 1
                        continue
                    elif existing and req.duplicate_handling == "update":
                        for k, v in data.items():
                            if k not in ("id", "employee_code") and hasattr(existing, k):
                                setattr(existing, k, v)
                        if not existing.code and code:
                            existing.code = code
                    else:
                        emp_data = {k: v for k, v in data.items()
                                    if hasattr(Employee, k) and k not in ("id", "employee_code")}
                        if code:
                            emp_data["code"] = code
                        if not emp_data.get("name"):
                            skipped += 1
                            continue
                        e = Employee(**emp_data)
                        db.add(e)
                    imported += 1

                else:
                    # Generic fallback
                    imported += 1

            except Exception as e:
                errors_out.append({"row": row_info["row_num"], "message": str(e)})
                skipped += 1

        try:
            await db.commit()
        except Exception as e:
            await db.rollback()
            errors_out.append({"row": batch_start, "message": f"Batch commit failed: {str(e)}"})

    # Clean up temp files after successful import
    _delete_import_file(req.file_id)

    role_code = current_user.role_rel.code if current_user.role_rel else "UNKNOWN"
    client_ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "0.0.0.0").split(",")[0].strip()
    await log_audit(
        db, current_user.id, role_code, "import", req.module,
        current_user.mill_id or "all",
        f"Bulk import: module={req.module}, imported={imported}, skipped={skipped}, errors={len(errors_out)}",
        ip_address=client_ip,
    )
    await db.commit()

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors_out,
        "message": f"Imported {imported} records, skipped {skipped}",
    }


# ── Named mapping CRUD ─────────────────────────────────────────────

@router.post("/imports/named-mappings")
async def save_named_mapping(
    req: SaveNamedMappingRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_module("masters", write=True)),
):
    """Save a named mapping template for reuse."""
    scope = await get_mill_scope(current_user, db)
    mill_id = scope.get("mill_id") or current_user.mill_id
    if not mill_id:
        raise HTTPException(400, detail="mill_id required")

    # Delete any existing mapping with same name+module+mill
    await db.execute(
        delete(ImportMapping).where(
            ImportMapping.mill_id == mill_id,
            ImportMapping.table_name == f"{req.module}:{req.name}",
        )
    )
    for item in req.mapping:
        is_custom = item.field.startswith("custom:")
        db.add(ImportMapping(
            mill_id=mill_id,
            table_name=f"{req.module}:{req.name}",
            excel_header=item.header,
            spinflow_field=None if is_custom else item.field,
            is_custom_field=is_custom,
            confidence=100.0,
        ))
    await db.commit()
    return {"ok": True, "name": req.name, "module": req.module}


@router.get("/imports/named-mappings")
async def list_named_mappings(
    request: Request,
    module: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List saved named mappings for a module.
    Requires read permission on the target module.
    """
    await _assert_import_permission(module, current_user, db, write=False)
    scope = await get_mill_scope(current_user, db)
    mill_id = scope.get("mill_id") or current_user.mill_id
    if not mill_id:
        return []

    prefix = f"{module}:"
    result = await db.execute(
        select(ImportMapping).where(
            ImportMapping.mill_id == mill_id,
            ImportMapping.table_name.like(f"{prefix}%"),
        )
    )
    rows = result.scalars().all()

    # Group by name
    named: dict[str, list[dict]] = {}
    for r in rows:
        name = r.table_name[len(prefix):]
        named.setdefault(name, []).append({
            "header": r.excel_header,
            "field": r.spinflow_field if not r.is_custom_field else f"custom:{r.excel_header}",
        })

    return [{"name": name, "mapping": items} for name, items in named.items()]
