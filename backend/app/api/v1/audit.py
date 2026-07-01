"""Enhanced audit log API — Wave 4A.

Changes vs. original:
  - New filters: category, severity, mill_id, module
  - Scoping: uses AuditLog.company_id / mill_id columns (set by enhanced log_audit)
  - Soft-delete:  DELETE /audit/logs/{id}
  - Hard-delete:  DELETE /audit/logs/{id}?hard=true  (SUPER_ADMIN only)
  - Export:       GET    /audit/logs/export?format=csv|xlsx
  - Hides soft-deleted rows by default (include_deleted=false)
  - Expanded response model with Wave 4A fields
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, func, or_, update
from typing import Optional
from pydantic import BaseModel
from datetime import datetime, date
import io
import csv

from app.db.session import get_db
from app.core.deps import require_module, get_mill_scope
from app.models.user import User
from app.models.audit import AuditLog

router = APIRouter()


# ---------------------------------------------------------------------------
# Pydantic response model
# ---------------------------------------------------------------------------

class AuditLogResponse(BaseModel):
    id: str
    timestamp: str
    user_name: Optional[str] = None
    role: Optional[str] = None
    action: str
    entity: str
    entity_id: Optional[str] = None
    entity_name: Optional[str] = None
    details: Optional[str] = None
    old_value: Optional[str] = None
    new_value: Optional[str] = None
    ip_address: Optional[str] = None
    category: Optional[str] = None
    severity: Optional[str] = None
    company_id: Optional[str] = None
    company_name: Optional[str] = None
    mill_id: Optional[str] = None
    mill_name: Optional[str] = None
    module: Optional[str] = None
    deleted_at: Optional[str] = None

    class Config:
        from_attributes = True


def _to_response(log: AuditLog) -> AuditLogResponse:
    return AuditLogResponse(
        id=log.id,
        timestamp=log.created_at.isoformat() if log.created_at else "",
        user_name=log.user_name or "System",
        role=log.role,
        action=log.action,
        entity=log.entity,
        entity_id=log.entity_id,
        entity_name=log.entity_name,
        details=log.details,
        old_value=log.old_value,
        new_value=log.new_value,
        ip_address=log.ip_address,
        category=log.category,
        severity=log.severity,
        company_id=log.company_id,
        company_name=log.company_name,
        mill_id=log.mill_id,
        mill_name=log.mill_name,
        module=log.module,
        deleted_at=log.deleted_at.isoformat() if log.deleted_at else None,
    )


# ---------------------------------------------------------------------------
# Shared filter builder
# ---------------------------------------------------------------------------

def _build_query(
    *,
    scope: dict,
    action: Optional[str],
    entity: Optional[str],
    entity_id: Optional[str],
    user_id: Optional[str],
    company_id: Optional[str],
    mill_id: Optional[str],
    category: Optional[str],
    severity: Optional[str],
    module: Optional[str],
    date_from: Optional[date],
    date_to: Optional[date],
    search: Optional[str],
    include_deleted: bool,
):
    query = select(AuditLog).where(AuditLog.action.isnot(None))

    # Hide soft-deleted rows by default
    if not include_deleted:
        query = query.where(AuditLog.deleted_at.is_(None))

    # Tenant scope — prefer direct company_id/mill_id cols (Wave 4A); fall back to
    # user subquery for older rows that pre-date those columns.
    scope_mill = scope.get("mill_id")
    scope_company = scope.get("company_id")

    if scope_mill:
        query = query.where(
            or_(
                AuditLog.mill_id == scope_mill,
                AuditLog.user_id.in_(
                    select(User.id).where(User.mill_id == scope_mill)
                ),
            )
        )
    elif scope_company:
        query = query.where(
            or_(
                AuditLog.company_id == scope_company,
                AuditLog.user_id.in_(
                    select(User.id).where(User.company_id == scope_company)
                ),
            )
        )

    # Explicit filters
    if action:
        query = query.where(AuditLog.action == action)
    if entity:
        query = query.where(AuditLog.entity == entity)
    if entity_id:
        query = query.where(AuditLog.entity_id == entity_id)
    if user_id:
        query = query.where(AuditLog.user_id == user_id)
    if company_id:
        query = query.where(
            or_(
                AuditLog.company_id == company_id,
                AuditLog.user_id.in_(
                    select(User.id).where(User.company_id == company_id)
                ),
            )
        )
    if mill_id:
        query = query.where(
            or_(
                AuditLog.mill_id == mill_id,
                AuditLog.user_id.in_(
                    select(User.id).where(User.mill_id == mill_id)
                ),
            )
        )
    if category:
        query = query.where(AuditLog.category == category)
    if severity:
        query = query.where(AuditLog.severity == severity)
    if module:
        query = query.where(AuditLog.module == module)
    if date_from:
        query = query.where(
            AuditLog.created_at >= datetime.combine(date_from, datetime.min.time())
        )
    if date_to:
        query = query.where(
            AuditLog.created_at <= datetime.combine(date_to, datetime.max.time())
        )
    if search:
        query = query.where(
            or_(
                AuditLog.details.ilike(f"%{search}%"),
                AuditLog.user_name.ilike(f"%{search}%"),
                AuditLog.entity.ilike(f"%{search}%"),
                AuditLog.entity_name.ilike(f"%{search}%"),
            )
        )

    return query


# ---------------------------------------------------------------------------
# GET /audit/logs
# ---------------------------------------------------------------------------

@router.get("/audit/logs")
async def get_audit_logs(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_module("audit")),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=1000),
    action: Optional[str] = None,
    entity: Optional[str] = None,
    entity_id: Optional[str] = None,
    user_id: Optional[str] = None,
    company_id: Optional[str] = None,
    mill_id: Optional[str] = None,
    category: Optional[str] = None,
    severity: Optional[str] = None,
    module: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    search: Optional[str] = None,
    include_deleted: bool = False,
):
    scope = await get_mill_scope(current_user, db)
    query = _build_query(
        scope=scope,
        action=action, entity=entity, entity_id=entity_id,
        user_id=user_id, company_id=company_id, mill_id=mill_id,
        category=category, severity=severity, module=module,
        date_from=date_from, date_to=date_to, search=search,
        include_deleted=include_deleted,
    )

    query = query.order_by(desc(AuditLog.created_at))

    total = (await db.execute(
        select(func.count()).select_from(query.subquery())
    )).scalar() or 0

    query = query.offset((page - 1) * page_size).limit(page_size)
    logs = (await db.execute(query)).scalars().all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size if page_size > 0 else 0,
        "data": [_to_response(log) for log in logs],
    }


# ---------------------------------------------------------------------------
# GET /audit/logs/export
# ---------------------------------------------------------------------------

@router.get("/audit/logs/export")
async def export_audit_logs(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_module("audit")),
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    action: Optional[str] = None,
    entity: Optional[str] = None,
    entity_id: Optional[str] = None,
    user_id: Optional[str] = None,
    company_id: Optional[str] = None,
    mill_id: Optional[str] = None,
    category: Optional[str] = None,
    severity: Optional[str] = None,
    module: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    search: Optional[str] = None,
    include_deleted: bool = False,
):
    scope = await get_mill_scope(current_user, db)
    query = _build_query(
        scope=scope,
        action=action, entity=entity, entity_id=entity_id,
        user_id=user_id, company_id=company_id, mill_id=mill_id,
        category=category, severity=severity, module=module,
        date_from=date_from, date_to=date_to, search=search,
        include_deleted=include_deleted,
    )
    query = query.order_by(desc(AuditLog.created_at)).limit(10000)
    logs = (await db.execute(query)).scalars().all()

    headers = [
        "timestamp", "user_name", "role", "action", "entity", "entity_name",
        "entity_id", "category", "severity", "module",
        "company_name", "mill_name", "details", "ip_address",
    ]

    def _row(log: AuditLog) -> list:
        return [
            log.created_at.isoformat() if log.created_at else "",
            log.user_name or "System",
            log.role or "",
            log.action,
            log.entity,
            log.entity_name or "",
            log.entity_id or "",
            log.category or "",
            log.severity or "",
            log.module or "",
            log.company_name or "",
            log.mill_name or "",
            log.details or "",
            log.ip_address or "",
        ]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for log in logs:
            writer.writerow(_row(log))
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=audit_logs.csv"},
        )

    # xlsx
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for log in logs:
        ws.append(_row(log))

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=audit_logs.xlsx"},
    )


# ---------------------------------------------------------------------------
# DELETE /audit/logs/{log_id}  (soft or hard)
# ---------------------------------------------------------------------------

@router.delete("/audit/logs/{log_id}", status_code=200)
async def delete_audit_log(
    log_id: str,
    hard: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_module("audit")),
):
    role_code = current_user.role_rel.code if current_user.role_rel else (current_user.role or "")

    if hard and role_code != "SUPER_ADMIN":
        raise HTTPException(status_code=403, detail="Hard-delete requires SUPER_ADMIN role")

    log = await db.get(AuditLog, log_id)
    if not log:
        raise HTTPException(status_code=404, detail="Audit log not found")

    # Scope check — non-SUPER_ADMIN can only delete within their company
    if role_code != "SUPER_ADMIN":
        scope = await get_mill_scope(current_user, db)
        sc = scope.get("company_id")
        if sc and log.company_id and log.company_id != sc:
            raise HTTPException(status_code=403, detail="Cannot delete logs outside your company")

    if hard:
        await db.delete(log)
    else:
        await db.execute(
            update(AuditLog)
            .where(AuditLog.id == log_id)
            .values(
                deleted_at=datetime.utcnow(),
                deleted_by=current_user.id,
            )
        )

    await db.commit()
    return {"success": True, "hard": hard, "id": log_id}
