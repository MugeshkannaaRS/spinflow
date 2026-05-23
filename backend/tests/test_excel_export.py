from app.services.excel_export import production_report, payroll_report, gst_report
from openpyxl import load_workbook


def test_production_report_returns_xlsx():
    data = [
        {"date": "2026-05-01", "shift": "A", "machine": "MCH-001", "produced_kg": "1000", "status": "approved"},
    ]
    buf = production_report(data)
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.title == "Production Report"
    assert ws.cell(1, 1).value == "date"
    assert ws.cell(2, 1).value == "2026-05-01"
    wb.close()


def test_production_report_empty():
    buf = production_report([])
    wb = load_workbook(buf)
    assert wb.active.cell(1, 1).value == "No data"
    wb.close()


def test_payroll_report_returns_xlsx():
    data = [
        {"employee_name": "Kumar", "department": "Spinning", "gross_wage": "5750", "net_wage": "5106.88"},
    ]
    buf = payroll_report(data)
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.title == "Payroll Report"
    assert ws.cell(1, 1).value == "employee_name"
    assert ws.cell(2, 1).value == "Kumar"
    wb.close()


def test_gst_report_returns_xlsx():
    output_gst = {"cgst": 5000.0, "sgst": 5000.0, "igst": 0.0, "total": 10000.0}
    input_gst = {"total": 3000.0}
    buf = gst_report(output_gst, input_gst, 7000.0, 5, 2026)
    wb = load_workbook(buf)
    ws = wb.active
    assert "GST" in ws.title
    assert ws.cell(2, 1).value is not None
    wb.close()
