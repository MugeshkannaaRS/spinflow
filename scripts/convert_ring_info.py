#!/usr/bin/env python3
"""
convert_ring_info.py — turn the "Ring information sheet" (per-model Ring Frame
maintenance programmes with lubricant details) into the clean PM-schedule seed
format that SpinFlow's importer accepts.

Each sheet = one Ring Frame model (G5/2, G-30, Rx-240, G-33). Layout:
  rows 0-5  : company / address / title / line / date / model banner
  header    : S/L NO | Work description | Frequency of work | Lubricating Name/Brand
              | Quantity | Unit | Number of point | Per machine | Remarks
  data rows : the PM tasks
Ditto marks (") in the lubricant column mean "same as the row above".

Output columns (importable via Maintenance -> PM Schedules -> Import Schedule):
  Machine Code | Department | SL No | Work Description | Frequency | Frequency Days
  | Lubricant Name | Lubricant Quantity

USAGE
  python scripts/convert_ring_info.py "Ring information sheet.xlsx" -o Ring_PM_import.xlsx
"""
import argparse
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")


FREQ = {
    "day": 1, "days": 1, "week": 7, "weeks": 7,
    "month": 30, "months": 30, "year": 365, "years": 365,
}


def freq_to_days(label: str):
    if not label:
        return ""
    s = re.sub(r"\s+", " ", str(label).strip().lower())
    m = re.match(r"0?(\d+)\s*(day|days|week|weeks|month|months|year|years)", s)
    if m:
        return int(m.group(1)) * FREQ.get(m.group(2), 1)
    return ""


def cell(r, i):
    return str(r[i]).strip() if i < len(r) and r[i] is not None else ""


def is_ditto(v: str) -> bool:
    return v.strip() in ('"', "''", "”", "“", "‘‘", "’’")


def machine_code_from(sheet_name: str, banner: str) -> str:
    """Derive a compact machine code, e.g. 'Ring Frame REITER  G 5/2' -> 'Ring_G5-2'."""
    src = banner or sheet_name
    m = re.search(r"(G\s*-?\s*\d+\s*/?\s*\d*|Rx\s*-?\s*\d+)", src, re.I)
    token = (m.group(0) if m else sheet_name).strip()
    token = re.sub(r"\s+", "", token).replace("/", "-")
    return f"Ring_{token}"[:50]


def convert(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # find header row (has "work description")
        hdr_idx = None
        banner = ""
        for i, row in enumerate(rows[:12]):
            joined = " ".join(cell(row, c).lower() for c in range(min(4, len(row))))
            if "work description" in joined:
                hdr_idx = i
                break
            # model banner like "Ring Frame REITER G 5/2"
            if "ring frame" in cell(row, 0).lower():
                banner = cell(row, 0)
        if hdr_idx is None:
            continue
        mc = machine_code_from(sheet, banner)

        last_lub = ""
        last_unit = ""
        for row in rows[hdr_idx + 1:]:
            sl = cell(row, 0)
            desc = cell(row, 1)
            freq = cell(row, 2)
            lub = cell(row, 3)
            qty = cell(row, 4)
            unit = cell(row, 5)
            if not re.match(r"^\d+$", sl) or not desc:
                continue
            # carry lubricant name forward on ditto marks
            if is_ditto(lub) or not lub:
                lub_name = last_lub
            else:
                lub_name = lub
                last_lub = lub
            # carry unit forward on ditto marks
            if is_ditto(unit) or not unit:
                unit = last_unit
            else:
                last_unit = unit
            lub_qty = (f"{qty} {unit}".strip() if qty else "").strip()
            out.append([
                mc, "Ringframe", int(sl), desc, freq, freq_to_days(freq),
                lub_name or "", lub_qty,
            ])
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("-o", "--output", default="Ring_PM_import.xlsx")
    args = ap.parse_args()

    rows = convert(args.input)
    if not rows:
        sys.exit("No task rows extracted — check the file layout.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PM Schedule"
    ws.append([
        "Machine Code", "Department", "SL No", "Work Description",
        "Frequency", "Frequency Days", "Lubricant Name", "Lubricant Quantity",
    ])
    for r in rows:
        ws.append(r)
    wb.save(args.output)

    from collections import Counter
    machines = Counter(r[0] for r in rows)
    print(f"✓ Wrote {len(rows)} rows → {args.output}")
    print(f"  {len(machines)} machines: {dict(machines)}")
    print(f"  rows with lubricant: {sum(1 for r in rows if r[6])}")


if __name__ == "__main__":
    main()
