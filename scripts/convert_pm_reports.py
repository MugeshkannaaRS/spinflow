#!/usr/bin/env python3
"""
convert_pm_reports.py — reshape raw AACSL/AAYML maintenance reports into the
clean PM-schedule seed format that SpinFlow's importer expects.

The mill's original spreadsheets are human-formatted reports, not data tables:
  * rows 0-2  = company name / address / programme title  (skip)
  * a MACHINE BANNER row like "BLENDOMAT (BDT-019)" — machine code in parens,
    and sometimes "4 persons" / "24 machines" in trailing cells
  * a HEADER row: "S/L NO | Work description | Frequency of work |
                   Lubricating Name/Brand | Quantity | Remarks"
  * data rows until the next machine banner
  * one SHEET per department; sheet-name prefix "AA"/"CSL" = which mill

This script flattens all of that into a single sheet with columns the
DirectImportModal recognizes:
  Machine Code | Department | SL No | Work Description | Frequency |
  Frequency Days | Lubricant Name | Lubricant Quantity | Manpower Count |
  Machine Count

USAGE
-----
  python scripts/convert_pm_reports.py INPUT1.xlsx [INPUT2.xlsx ...] -o PM_Schedule_clean.xlsx
  # filter to one mill prefix (e.g. only CSL sheets):
  python scripts/convert_pm_reports.py "AACSL Workdescription.xlsx" --mill CSL -o csl_pm.xlsx
"""
import argparse
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")


# Frequency label → days (mirrors backend _FREQ_MAP).
FREQ_DAYS = {
    "daily": 1, "weekly": 7, "7 days": 7, "fortnightly": 14,
    "01 month": 30, "1 month": 30, "monthly": 30,
    "02 month": 60, "2 month": 60,
    "03 month": 90, "3 month": 90, "quarterly": 90,
    "04 month": 120, "4 month": 120,
    "06 month": 180, "6 month": 180, "06 months": 180, "6 months": 180,
    "01 year": 365, "1 year": 365, "yearly": 365, "annually": 365,
    "02 years": 730, "2 years": 730, "2.5 year": 912,
    "03 years": 1095, "3 years": 1095, "04 years": 1460, "05 years": 1825,
}

HEADER_TOKENS = {"s/l", "sl", "work description", "work", "frequency"}
OUT_HEADERS = [
    "Machine Code", "Department", "SL No", "Work Description", "Frequency",
    "Frequency Days", "Lubricant Name", "Lubricant Quantity",
    "Manpower Count", "Machine Count",
]


def freq_to_days(label: str):
    if not label:
        return ""
    s = re.sub(r"\s+", " ", str(label).strip().lower())
    if s in FREQ_DAYS:
        return FREQ_DAYS[s]
    m = re.match(r"(\d+)\s*day", s)
    if m:
        return int(m.group(1))
    m = re.match(r"0?(\d+)\s*month", s)
    if m:
        return int(m.group(1)) * 30
    m = re.match(r"0?(\d+)\s*year", s)
    if m:
        return int(m.group(1)) * 365
    return ""


def clean(v):
    return re.sub(r"\s+", " ", str(v).strip()) if v is not None else ""


def parse_int(v):
    m = re.search(r"\d+", str(v or ""))
    return int(m.group()) if m else ""


def is_header_row(cells):
    """A column-header row: has a description/activities token AND a frequency
    token somewhere in the first few cells. Tolerates typos ('FREQUANCY',
    'ACTIVITIES') and newlines inside header cells."""
    joined = " ".join(c.lower().replace("\n", " ") for c in cells[:5] if c)
    has_desc = ("work description" in joined) or ("activities" in joined) or ("activity" in joined)
    has_freq = ("frequency" in joined) or ("frequancy" in joined) or ("frequ" in joined)
    sl_first = cells and cells[0] and re.sub(r"[^a-z]", "", cells[0].lower()) in ("slno", "sl", "slno.")
    return (has_desc and has_freq) or (sl_first and has_freq)


def is_banner_candidate(cells):
    """A machine-banner row: text in col 0, no S/L digit, not a header, not
    title noise. Trailing cells may hold manpower/machine counts."""
    if not cells or not cells[0]:
        return False
    low = cells[0].lower()
    if any(tok in low for tok in TITLE_NOISE):
        return False
    if is_header_row(cells):
        return False
    # First cell should be a label, not a numbered task row
    if re.match(r"^\d+$", cells[0].strip()):
        return False
    return True


TITLE_NOISE = ("maintenance programe", "maintenance program", "general work",
               "limited", "gazipur", "sreepur", "nagar", "date:")


def _manpower_machines(cells):
    manpower = machines = ""
    for c in cells:
        cl = c.lower()
        if re.search(r"person|persom|\bman\b", cl):
            manpower = parse_int(c)
        elif re.search(r"machine|\bmcs?\b", cl):
            machines = parse_int(c)
    return manpower, machines


def machine_from_banner(cells):
    """Interpret a machine-banner row.

    A banner is the row directly above the 'S/L NO | Work description ...'
    header. It holds the machine name, optionally with code(s) in parens, plus
    optional 'N persons' / 'N machines' enrichment in trailing cells. Returns
    (code, manpower, machines) or None if the row is title/noise.
    """
    first = cells[0] if cells else ""
    if not first:
        return None
    low = first.lower()
    if any(tok in low for tok in TITLE_NOISE):
        return None
    # Prefer an explicit code in parentheses that looks like a machine code
    # (short, contains a digit or hyphen). Multi-code parens → take the first.
    m = re.search(r"\(([^)]+)\)", first)
    code = None
    if m:
        inner = m.group(1).strip()
        first_code = re.split(r"[,/]", inner)[0].strip()
        if re.search(r"[0-9\-]", first_code) and len(first_code) <= 20:
            code = first_code.replace(" ", "_")
    if not code:
        # No usable parens — use the machine NAME itself as the code,
        # cleaned to a compact token (e.g. "MURATA 7V-II - AUTOCONE..." ->
        # "MURATA_7V-II"). Take the portion before the first dash/slash.
        name = re.split(r"[-/]", first)[0].strip()
        if not name or len(name) < 2:
            return None
        code = re.sub(r"\s+", "_", name)[:30]
    manpower, machines = _manpower_machines(cells[1:])
    return code, manpower, machines


def convert(paths, mill_prefix=None):
    out_rows = []
    for path in paths:
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            print(f"  ! skip {path}: {e}", file=sys.stderr)
            continue
        for sheet in wb.sheetnames:
            if mill_prefix and not sheet.upper().startswith(mill_prefix.upper()):
                continue
            ws = wb[sheet]
            # Department = sheet name minus mill prefix
            dept = re.sub(r"^(AA|CSL|AAYML|AACSL)\s*", "", sheet, flags=re.I).strip() or sheet

            # Materialize the sheet so we can look ahead (a banner is confirmed
            # only when the next non-empty row is a header).
            grid = [[clean(c) for c in row] for row in ws.iter_rows(values_only=True)]
            non_empty = [(i, cells) for i, cells in enumerate(grid) if any(cells)]

            cur_machine = cur_manpower = cur_machines = ""
            in_data = False
            for pos, (i, cells) in enumerate(non_empty):
                # Header row → start of a data block for the current machine
                if is_header_row(cells):
                    in_data = True
                    continue
                # Banner candidate confirmed if the NEXT non-empty row is a header
                nxt = non_empty[pos + 1][1] if pos + 1 < len(non_empty) else None
                if is_banner_candidate(cells) and nxt is not None and is_header_row(nxt):
                    parsed = machine_from_banner(cells)
                    if parsed:
                        cur_machine, cur_manpower, cur_machines = parsed
                        in_data = False
                        continue
                if not in_data or not cur_machine:
                    continue
                # Data row: [sl, work desc, freq, lubricant, qty, remarks]
                sl = parse_int(cells[0]) if len(cells) > 0 else ""
                desc = cells[1] if len(cells) > 1 else ""
                freq = cells[2] if len(cells) > 2 else ""
                lub = cells[3] if len(cells) > 3 else ""
                qty = cells[4] if len(cells) > 4 else ""
                if not desc:
                    continue
                out_rows.append([
                    cur_machine, dept, sl, desc, freq, freq_to_days(freq),
                    lub, qty, cur_manpower, cur_machines,
                ])
    return out_rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inputs", nargs="+", help="raw report .xlsx files")
    ap.add_argument("-o", "--output", default="PM_Schedule_clean.xlsx")
    ap.add_argument("--mill", help="only sheets whose name starts with this prefix (AA / CSL)")
    args = ap.parse_args()

    rows = convert(args.inputs, args.mill)
    if not rows:
        sys.exit("No data rows extracted — check the input files / --mill filter.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PM Schedule"
    ws.append(OUT_HEADERS)
    for r in rows:
        ws.append(r)
    wb.save(args.output)
    machines = sorted({r[0] for r in rows})
    depts = sorted({r[1] for r in rows})
    print(f"✓ Wrote {len(rows)} rows → {args.output}")
    print(f"  {len(machines)} machines across {len(depts)} departments: {', '.join(depts)}")


if __name__ == "__main__":
    main()
